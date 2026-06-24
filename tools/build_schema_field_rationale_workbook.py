from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


OUTPUT = Path("/Users/aminsuliman/Downloads/MPG_database_H2/output/query_optimized_schema_field_rationale.xlsx")


HEADERS = ["Table", "Field", "What is it?", "Why is it here?", "How it helps querying/storage", "Recommendation"]


BLUE = "1F77B4"
GREEN = "22C55E"
PURPLE = "7C3AED"
GRAY = "6B7280"
RED = "B91C1C"


def add_sheet(wb, title, headers, rows, color):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor=color)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    widths = [30, 32, 45, 52, 58, 30]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 45 if i > 1 else 30
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb,
        "Overview",
        ["Section", "Recommendation", "Why it matters"],
        [
            ["Goal", "Optimize schema for fast filtering/querying while preserving flexible annotations.", "Common filters become indexed columns; variable fields go to annotation tables."],
            ["Core rule", "Use real columns for fields clinicians filter frequently.", "This avoids repeated joins, string parsing, and slow wildcard search."],
            ["Flexible rule", "Use annotation tables for external, assay-specific, or rarely queried fields.", "This prevents schema bloat while keeping values queryable."],
            ["Raw/provenance rule", "Keep raw text only where it adds traceability and avoid duplicating it on every child row.", "This matters for storage size with large VCF/CNV datasets."],
            ["SNV/indel model", "Use small_variants, small_variant_sample_calls, small_variant_annotations, and small_variant_external_annotations.", "Separates variant identity, sample evidence, transcript effect, and external evidence."],
            ["CNV/SV model", "Use genomic_segments, segment_annotations, genomic_links, karyotype, and notes.", "Keeps core interval filters fast while preserving flexible source metadata."],
            ["Excluded from this pass", "variant_classifications and signed_out_calls are intentionally excluded.", "The current question is query/filter schema, not clinical sign-out workflow."],
        ],
        GRAY,
    )

    table_data = {
        "small_variants": [
            ["small_variants", "small_variant_id", "Primary key for the normalized variant row.", "Gives every unique variant a stable internal ID for joins.", "Integer primary keys make joins fast and compact.", "Keep"],
            ["small_variants", "chromosome", "Chromosome containing the variant.", "Required for genomic location filtering.", "Index with genome_build and position for fast location search.", "Keep"],
            ["small_variants", "position", "1-based VCF position/start coordinate.", "Needed for exact position and hotspot searches.", "Indexed location queries avoid scanning annotation text.", "Keep"],
            ["small_variants", "end_position", "End coordinate for indels or ranged small variants.", "Position alone is not always enough for indels.", "Supports range display/query without deriving every time.", "Add/Keep"],
            ["small_variants", "variant_id", "Original VCF ID field.", "Preserves what the source file said.", "Useful for display/provenance; index if searched often.", "Keep"],
            ["small_variants", "rs_id", "Parsed dbSNP rs identifier when present.", "Clinicians may search known rsIDs.", "Dedicated index is faster than searching generic external annotations.", "Keep"],
            ["small_variants", "ref_allele", "Reference allele.", "Needed for exact variant identity.", "Part of normalized identity; prevents coordinate-only false matches.", "Keep"],
            ["small_variants", "alt_allele", "Alternate allele.", "Needed for exact variant identity.", "Part of normalized identity and multi-allelic tracking.", "Keep"],
            ["small_variants", "variant_type", "SNV, insertion, deletion, delins, MNV, complex, unknown.", "Clinicians filter different variant classes.", "Low-cardinality indexed column filters quickly.", "Keep"],
            ["small_variants", "genome_build", "Reference build such as GRCh37 or GRCh38.", "Coordinates and alleles cannot be interpreted without build.", "Index with chromosome/position; prevents cross-build false matches.", "Keep"],
            ["small_variants", "normalized_key", "Canonical deduplication key, e.g. build|chr|pos|ref|alt.", "Prevents duplicate variant identity rows across imports.", "Unique/indexed lookup and deduplication.", "Keep/Add"],
        ],
        "small_variant_sample_calls": [
            ["small_variant_sample_calls", "small_variant_call_id", "Primary key for sample-specific variant call.", "A variant can be observed in many sample test results.", "Compact join key for sample-level evidence.", "Keep"],
            ["small_variant_sample_calls", "small_variant_id", "Foreign key to small_variants.", "Links sample evidence to variant identity.", "Indexed FK supports fast joins from variants to calls.", "Keep"],
            ["small_variant_sample_calls", "sample_test_result_id", "Foreign key to the imported test/result.", "Connects call to sample, source file, protocol, and pipeline.", "Indexed field supports case/result filtering.", "Keep"],
            ["small_variant_sample_calls", "qual", "VCF QUAL score.", "Supports caller confidence/QC display.", "Numeric filter/sort for quality review.", "Keep"],
            ["small_variant_sample_calls", "filter_status", "VCF FILTER value such as PASS or LowQual.", "Common QC filter.", "Indexed low-cardinality field makes PASS-only queries fast.", "Keep"],
            ["small_variant_sample_calls", "genotype", "Raw genotype string such as 0/1 or 1/1.", "Preserves exact caller genotype.", "Useful for technical display/filtering.", "Keep"],
            ["small_variant_sample_calls", "zygosity", "User-friendly genotype category.", "Clinicians search heterozygous/homozygous/hemizygous/mosaic.", "Avoids parsing genotype during every query.", "Keep/Add"],
            ["small_variant_sample_calls", "phased", "Whether genotype is phased.", "Useful when phasing is available.", "Boolean filter with small storage cost.", "Keep"],
            ["small_variant_sample_calls", "ref_depth", "Reference allele read depth.", "Shows read evidence behind the call.", "Numeric QC filter and VAF validation.", "Keep"],
            ["small_variant_sample_calls", "alt_depth", "Alternate allele read depth.", "Shows read evidence behind the call.", "Numeric QC filter and VAF validation.", "Keep"],
            ["small_variant_sample_calls", "total_depth", "Total depth, usually FORMAT DP.", "Read depth is a V1 QC filter.", "Direct numeric index/filter; not always exactly ref+alt.", "Keep"],
            ["small_variant_sample_calls", "genotype_quality", "FORMAT GQ when present.", "Supports genotype confidence review.", "Numeric quality filter without parsing sample_values.", "Keep"],
            ["small_variant_sample_calls", "variant_allele_fraction", "VAF calculated/imported as decimal.", "Filtering spreadsheet calls out VAF as important.", "Direct numeric filtering is faster than calculating from text.", "Keep; use instead of duplicate allele_balance"],
            ["small_variant_sample_calls", "format_keys", "Raw FORMAT key list from VCF.", "Preserves caller-specific sample fields.", "Hidden provenance; avoids losing unsupported fields.", "Keep hidden"],
            ["small_variant_sample_calls", "sample_values", "Raw sample value string from VCF.", "Preserves exact source sample call.", "Hidden provenance/debug; not a main query field.", "Keep hidden"],
            ["small_variant_sample_calls", "info_raw", "Original INFO field.", "Provides raw VCF context without repeating ANN raw per annotation.", "Centralized provenance at call/variant line level.", "Keep hidden"],
            ["small_variant_sample_calls", "raw_vcf_line", "Full original VCF record line.", "Allows reconstruction/debugging of parsed values.", "Avoids needing annotation_raw on every annotation row.", "Keep hidden"],
            ["small_variant_sample_calls", "line_number", "Source VCF line number.", "Helps debug import issues and trace to file.", "Small field; useful for troubleshooting.", "Keep hidden"],
        ],
        "small_variant_annotations": [
            ["small_variant_annotations", "small_variant_annotation_id", "Primary key for transcript/gene annotation row.", "One variant can have multiple transcript annotations.", "Compact join key for annotation-level rows.", "Keep"],
            ["small_variant_annotations", "small_variant_id", "Foreign key to small_variants.", "Links transcript/gene effect to the variant.", "Indexed FK makes variant-to-annotation joins fast.", "Keep"],
            ["small_variant_annotations", "gene", "Gene symbol such as TP53 or OR4F5.", "Gene is a primary clinician search field.", "Dedicated indexed column is faster than key-value self-joins.", "Keep"],
            ["small_variant_annotations", "gene_id", "Stable gene identifier such as Ensembl ID.", "Gene symbols can change or be ambiguous.", "Helpful for normalization; hide by default if not user-facing.", "Keep hidden"],
            ["small_variant_annotations", "transcript", "Transcript ID.", "HGVS/consequence depend on transcript.", "Indexed transcript filter supports preferred transcript workflows.", "Keep"],
            ["small_variant_annotations", "is_preferred_transcript", "Whether annotation is preferred/canonical for display/filtering.", "Filtering sheet asks to clarify preferred transcript behavior.", "Boolean filter avoids ambiguous all-transcript results.", "Keep"],
            ["small_variant_annotations", "consequence", "Molecular consequence such as missense_variant.", "Core prioritization filter.", "Dedicated indexed column avoids key-value joins.", "Keep"],
            ["small_variant_annotations", "impact", "Impact category such as HIGH/MODERATE/LOW.", "Common prioritization filter.", "Low-cardinality indexed field filters quickly.", "Keep"],
            ["small_variant_annotations", "hgvs_c", "HGVS cDNA notation.", "Clinicians search/report cDNA notation.", "Direct text lookup; no parsing raw ANN.", "Keep"],
            ["small_variant_annotations", "hgvs_p", "HGVS protein notation.", "Clinicians search/report protein notation.", "Direct text lookup; supports protein-level filters.", "Keep"],
            ["small_variant_annotations", "annotation_source", "Tool/source that produced transcript annotation.", "Useful provenance when SnpEff/VEP/ANNOVAR differ.", "Hide by default; helps audit without raw duplication.", "Keep hidden"],
            ["small_variant_annotations", "annotation_version", "Version of annotation tool/database.", "Explains why gene/consequence may change over time.", "Provenance; not a main filter but useful for review.", "Keep hidden"],
            ["small_variant_annotations", "annotation_raw", "Original raw ANN subentry.", "Mostly duplicates raw_vcf_line/info_raw provenance.", "Not useful for fast querying and bloats storage.", "Remove/deprecate"],
            ["small_variant_annotations", "is_primary_transcript", "Older/alternate flag for main transcript.", "Overlaps with is_preferred_transcript.", "Keeping both creates confusion; use one policy field.", "Remove or map to is_preferred_transcript"],
        ],
        "small_variant_external": [
            ["small_variant_external_annotations", "external_annotation_id", "Primary key for external annotation row.", "External sources can produce many values per variant.", "Compact join key.", "Keep"],
            ["small_variant_external_annotations", "small_variant_id", "Foreign key to small_variants.", "External evidence belongs to a variant identity.", "Indexed FK supports fast joins.", "Keep"],
            ["small_variant_external_annotations", "source_name", "External source name such as ClinVar, gnomAD, CADD.", "Users need to know where evidence came from.", "Index with annotation_name for source-specific searches.", "Keep"],
            ["small_variant_external_annotations", "source_version", "External source release/version.", "ClinVar/gnomAD values change over time.", "Advanced provenance; not usually primary filter.", "Keep hidden"],
            ["small_variant_external_annotations", "genome_build", "Build used by external annotation.", "External mapping can be build-specific.", "Prevents cross-build confusion; useful in source filtering.", "Keep"],
            ["small_variant_external_annotations", "external_variant_id", "Source-specific identifier such as ClinVar Variation ID.", "Links to source records; stores dbSNP/ClinVar/OMIM IDs if not primary rs_id.", "Text lookup for external IDs without changing schema.", "Keep"],
            ["small_variant_external_annotations", "annotation_type", "Broad category such as clinical_significance or population_frequency.", "Groups flexible external fields.", "Indexed category narrows queries before value filtering.", "Keep"],
            ["small_variant_external_annotations", "annotation_name", "Specific field name such as gnomAD_AF or CADD_PHRED.", "Allows many external fields without adding columns.", "Index with typed value columns for flexible fast lookup.", "Keep"],
            ["small_variant_external_annotations", "text_value", "Text value for external annotation.", "Stores ClinVar labels, OMIM diseases, domains, PMIDs.", "Indexed text lookup for exact external filters.", "Keep"],
            ["small_variant_external_annotations", "numeric_value", "Numeric value for external annotation.", "Stores AF, prediction scores, counts.", "Numeric index supports ranges such as AF < 0.01.", "Keep"],
            ["small_variant_external_annotations", "boolean_value", "Boolean external value.", "Stores flags such as external_evidence_available.", "Boolean filters are compact and fast.", "Keep"],
            ["small_variant_external_annotations", "value_type", "TEXT/NUMBER/BOOLEAN marker.", "Tells the app which value column to use.", "Avoids ambiguous typed-value interpretation.", "Keep"],
            ["small_variant_external_annotations", "population", "Population context for frequency.", "gnomAD AF needs population context.", "Allows filtering global vs subpopulation frequency.", "Keep"],
            ["small_variant_external_annotations", "subpopulation", "Specific subgroup such as AFR/EUR/ALL.", "Frequency meaning depends on ancestry group.", "Supports subpopulation-specific AF filtering.", "Keep"],
            ["small_variant_external_annotations", "retrieved_date", "Date external annotation was imported/refreshed.", "External databases change over time.", "Provenance; can support audits but not front-page filter.", "Keep hidden"],
            ["small_variant_external_annotations", "evidence_url", "Link to external evidence page/source.", "Users may need click-through to ClinVar, OMIM, PubMed, etc.", "Stores links without bloating core tables.", "Keep"],
            ["small_variant_external_annotations", "raw_value", "Original external value text.", "Preserves source-specific details not normalized.", "Hidden provenance; avoid querying this normally.", "Keep hidden"],
        ],
        "genomic_segments": [
            ["genomic_segments", "segment_id", "Primary key for CNV/SV segment.", "Each parsed segment needs a stable ID.", "Compact join key for annotations/links.", "Keep"],
            ["genomic_segments", "sample_test_result_id", "Foreign key to source test result.", "Connects segment to sample/source/pipeline.", "Indexed case/result filtering.", "Keep"],
            ["genomic_segments", "karyotype_id", "Optional link to karyotype context.", "Needed for ISCN/cytogenetic-derived segments.", "Keeps karyotype text separate from segment rows.", "Keep"],
            ["genomic_segments", "event_group_id", "Label/group for related event pieces.", "Groups pieces of complex events.", "Indexed grouping helps event review.", "Keep"],
            ["genomic_segments", "chromosome", "Chromosome for segment.", "Core CNV/SV location filter.", "Index with start/stop for interval queries.", "Keep"],
            ["genomic_segments", "start_pos", "Base-pair start coordinate.", "Core interval-overlap filter.", "Real column is faster than annotation join.", "Keep"],
            ["genomic_segments", "stop_pos", "Base-pair end coordinate.", "Core interval-overlap filter and size calculation.", "Real column is faster than annotation join.", "Keep"],
            ["genomic_segments", "event_size_bp", "Computed segment size in bp.", "Filtering sheet has event size as V1 filter.", "Avoids computing stop-start for every query; indexable.", "Add/Keep"],
            ["genomic_segments", "cytoband_start", "Start cytoband.", "Historical cytogenetic records use cytobands.", "Direct cytoband filtering without annotation join.", "Add/Keep"],
            ["genomic_segments", "cytoband_end", "End cytoband.", "Needed for cytoband interval display/filtering.", "Direct cytoband filtering without annotation join.", "Add/Keep"],
            ["genomic_segments", "event_type", "Deletion, duplication, gain, loss, LOH, etc.", "Clinicians filter structural-change class.", "Low-cardinality index supports fast filtering.", "Keep"],
            ["genomic_segments", "copy_number", "Observed/inferred copy number.", "Distinguishes losses/gains/amplifications.", "Numeric/indexed filter; better than storing in annotations.", "Keep"],
            ["genomic_segments", "genome_build", "Reference build for coordinates.", "Coordinates need build context.", "Index with interval coordinates.", "Keep"],
            ["genomic_segments", "confidence", "Parser/call confidence, not clinical significance.", "Filtering sheet asks for parser confidence/QC.", "Filter low-confidence parsed calls; clarify GUI label.", "Keep/rename label"],
            ["genomic_segments", "number_of_sites", "Array/probe/site support count.", "Useful support metric when common.", "If queried often, keep as core; otherwise could be annotation.", "Keep if common filter"],
            ["genomic_segments", "raw_segment_text", "Specific source text that produced this segment.", "Traceability without repeating full result text everywhere.", "Hidden provenance; better than generic annotations string.", "Keep"],
            ["genomic_segments", "ambiguity_flag", "Whether parsed call has ambiguity/uncertainty.", "User noted this is for ambiguous cases.", "Boolean filter for uncertain calls; details can go in annotation/note.", "Add/Keep"],
            ["genomic_segments", "line_number", "Source line number.", "Debugging and traceability.", "Small storage; useful when import rows fail or need audit.", "Keep hidden"],
            ["genomic_segments", "raw_iscn", "Old/ambiguous raw ISCN field.", "Can duplicate raw_segment_text or result-level source text.", "Avoid duplicate raw storage; use raw_segment_text instead.", "Deprecate unless distinct"],
            ["genomic_segments", "annotations", "Old pipe-delimited annotation storage.", "Legacy compatibility only.", "Bad for fast querying; replaced by segment_annotations.", "Deprecate"],
        ],
        "segment_annotations": [
            ["segment_annotations", "annotation_id", "Primary key for segment annotation row.", "Each extra field gets its own row.", "Compact join key.", "Keep"],
            ["segment_annotations", "segment_id", "Foreign key to genomic_segments.", "Links extra metadata to segment.", "Indexed FK supports fast joins.", "Keep"],
            ["segment_annotations", "annotation_name", "Name of source-specific extra field.", "Allows flexible metadata such as probe_count or BAF.", "Index with value columns for flexible lookup.", "Keep"],
            ["segment_annotations", "text_value", "Text annotation value.", "Stores string fields like array_platform.", "Indexed text search for exact key-value filters.", "Keep"],
            ["segment_annotations", "numeric_value", "Numeric annotation value.", "Stores probe count, LRR, BAF, scores.", "Numeric range filtering without parsing strings.", "Keep"],
            ["segment_annotations", "boolean_value", "Boolean annotation value.", "Stores true/false flags.", "Compact boolean filters.", "Keep"],
            ["segment_annotations", "value_type", "TEXT/NUMBER/BOOLEAN marker.", "Tells app which value column is meaningful.", "Avoids ambiguous values and supports typed UI filters.", "Keep"],
            ["segment_annotations", "source_column", "Original source column name.", "Preserves input provenance and typo/source labels.", "Does not require schema changes for new file formats.", "Keep"],
            ["segment_annotations", "ordinal_position", "Original source column order.", "Useful for export/display order.", "Small metadata cost; preserves file context.", "Add/Keep"],
        ],
        "genomic_links": [
            ["genomic_links", "link_id", "Primary key for segment relationship.", "Each real pair/link needs an ID.", "Compact key for link review.", "Keep"],
            ["genomic_links", "event_group_id", "Event group label or ID.", "Groups related structural-event pieces.", "Supports grouping and filtering complex events.", "Keep"],
            ["genomic_links", "source_seg_id", "Source segment in link.", "Represents one side of paired/breakpoint relationship.", "FK join to segment table.", "Keep"],
            ["genomic_links", "target_seg_id", "Target segment in link.", "Represents other side of paired relationship.", "FK join to segment table.", "Keep"],
            ["genomic_links", "link_type", "Type of link, e.g. translocation partner.", "Tells why two segments are connected.", "Indexed low-cardinality field supports relationship filtering.", "Keep"],
            ["genomic_links", "confidence", "Confidence of link parsing/evidence.", "Separates strong links from inferred links.", "Filter/review low-confidence links.", "Keep"],
        ],
        "karyotype": [
            ["karyotype", "karyotype_id", "Primary key for karyotype row.", "Segments may come from karyotype/ISCN context.", "Compact join key.", "Keep"],
            ["karyotype", "sample_test_result_id", "Foreign key to test result.", "Connects karyotype to sample/source.", "Indexed result filtering.", "Keep"],
            ["karyotype", "karyotype_text", "Original karyotype string/context.", "Preserves ISCN/cytogenetic result context.", "Hidden provenance/display field; not main filter.", "Keep"],
            ["karyotype", "cell_count", "Cell count associated with karyotype/clone.", "Important cytogenetic context when available.", "Numeric/context filter if needed.", "Keep"],
            ["karyotype", "parsed_status", "Parser status for karyotype text.", "Shows parsed/partial/unsupported state.", "Fast status filtering for review.", "Keep"],
            ["karyotype", "created_at", "Creation/import timestamp.", "Traceability.", "Sort/filter by import chronology if needed.", "Keep"],
        ],
        "notes": [
            ["notes", "note_id", "Primary key for note.", "Each note needs a stable ID.", "Compact lookup key.", "Keep"],
            ["notes", "target_table", "Which table the note attaches to.", "Allows notes on variants, segments, calls, results.", "Flexible without adding note columns to every table.", "Keep"],
            ["notes", "target_id", "ID of target row.", "Links note to specific object.", "Index with target_table for lookup.", "Keep"],
            ["notes", "note_type", "Category such as review_note or pathology_context.", "Helps filter/display note types.", "Low-cardinality filter.", "Keep"],
            ["notes", "note_text", "Free-text note.", "Captures human/context text not suited to structured columns.", "Not ideal for fast structured filtering; use for display/search only.", "Keep"],
            ["notes", "created_by", "Author/user/source.", "Audit trail.", "Display/provenance field.", "Keep"],
            ["notes", "created_at", "Timestamp.", "Audit trail and timeline.", "Sort/filter notes by date.", "Keep"],
        ],
    }

    all_field_rows = []
    for rows in table_data.values():
        all_field_rows.extend(rows)

    add_sheet(
        wb,
        "All fields rationale",
        HEADERS,
        all_field_rows,
        GRAY,
    )

    for name, rows in table_data.items():
        color = GREEN if name in {"genomic_segments", "segment_annotations", "genomic_links", "karyotype"} else PURPLE if name == "notes" else BLUE
        add_sheet(wb, name, HEADERS, rows, color)

    add_sheet(
        wb,
        "Recommended indexes",
        ["Table", "Index", "Columns", "Why it helps"],
        [
            ["small_variants", "idx_small_variants_location", "genome_build, chromosome, position", "Fast genomic location and hotspot search."],
            ["small_variants", "idx_small_variants_normalized", "normalized_key", "Fast deduplication and exact identity lookup."],
            ["small_variants", "idx_small_variants_rsid", "rs_id", "Fast dbSNP/rsID search."],
            ["small_variant_sample_calls", "idx_small_calls_result", "sample_test_result_id", "Fast case/result filtering."],
            ["small_variant_sample_calls", "idx_small_calls_filter", "filter_status", "Fast PASS/LowQual filtering."],
            ["small_variant_sample_calls", "idx_small_calls_zygosity", "zygosity", "Fast heterozygous/homozygous filtering."],
            ["small_variant_sample_calls", "idx_small_calls_vaf", "variant_allele_fraction", "Fast VAF range filtering."],
            ["small_variant_annotations", "idx_small_annotations_gene", "gene", "Fast gene search."],
            ["small_variant_annotations", "idx_small_annotations_consequence", "consequence", "Fast consequence filtering."],
            ["small_variant_annotations", "idx_small_annotations_impact", "impact", "Fast impact filtering."],
            ["small_variant_external_annotations", "idx_external_text", "source_name, annotation_name, text_value", "Fast ClinVar/OMIM/PubMed-style text lookups."],
            ["small_variant_external_annotations", "idx_external_numeric", "source_name, annotation_name, numeric_value", "Fast gnomAD AF / CADD / score range queries."],
            ["genomic_segments", "idx_segments_interval", "genome_build, chromosome, start_pos, stop_pos", "Fast CNV/SV interval overlap."],
            ["genomic_segments", "idx_segments_event_type", "event_type", "Fast deletion/duplication/LOH filtering."],
            ["genomic_segments", "idx_segments_size", "event_size_bp", "Fast CNV size filtering."],
            ["segment_annotations", "idx_segment_annotations_text", "annotation_name, text_value", "Fast flexible text metadata search."],
            ["segment_annotations", "idx_segment_annotations_numeric", "annotation_name, numeric_value", "Fast probe_count/score/range filtering."],
            ["genomic_links", "idx_genomic_links_group", "event_group_id", "Fast event group relationship review."],
        ],
        GRAY,
    )

    add_sheet(
        wb,
        "Deprecate or hide",
        ["Field", "Why repetitive/problematic", "Recommendation"],
        [
            ["small_variant_annotations.annotation_raw", "Duplicates raw_vcf_line/info_raw provenance.", "Remove/deprecate from optimized schema."],
            ["small_variant_sample_calls.allele_balance", "Repeats variant_allele_fraction concept.", "Use variant_allele_fraction as the real field and GUI label it VAF."],
            ["small_variant_annotations.is_primary_transcript", "Overlaps with is_preferred_transcript.", "Keep one preferred transcript flag."],
            ["genomic_segments.annotations", "Old pipe-delimited annotation field.", "Replace with segment_annotations for search."],
            ["sample_test_results.annotation_names", "Old shared-header design for pipe annotations.", "Stop using for new logic."],
            ["genomic_segments.raw_iscn", "Can duplicate raw_segment_text/source result text.", "Use only if it has a distinct segment-specific meaning; otherwise deprecate."],
        ],
        RED,
    )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
